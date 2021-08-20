from typing import List, Dict, Optional, Deque

from collections import deque

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook, ReadOnlyWorksheet

from .common import CVItem, CVItemFromIPTC, extract_code_from_string


def get_cv_items_from_iptc_xlsx(existing_topics: Dict[str, CVItem], filename: str) -> List[CVItemFromIPTC]:
    """Get MediaTopic CV items from an IPTC xlsx file"""

    spreadsheet = MediaTopicsSpreadsheet()
    spreadsheet.load_workbook_file(filename)
    spreadsheet.set_header_indices()
    spreadsheet.load_cv_items(existing_topics)

    return spreadsheet.rows


class MediaTopicsSpreadsheet:
    """Helper class to load MediaTopics from an xlsx spreadsheet"""

    workbook: Workbook
    worksheet: ReadOnlyWorksheet
    rows: List[CVItemFromIPTC]
    name_columns: Dict[str, int]
    subject_code_index: int
    wikidata_index: int
    _parent_qcode_stack: Deque[str]
    _parent_qcode_index: int

    def load_workbook_file(self, filename: str):
        """Load an xlsx workbook and select the active sheet (1st sheet)"""

        self.workbook = load_workbook(filename, read_only=True)
        self.worksheet = self.workbook.active

    def set_header_indices(self):
        """Iterates over sheet headers and sets the indices for the required columns"""

        self.name_columns = {}
        self.subject_code_index = -1
        self.wikidata_index = -1

        index: int = 0
        for cell in self.worksheet['2']:
            value: str = cell.value

            if value.startswith("Name ("):
                # This column contains translated Topic names
                language = value[6:-1].lower()
                self.name_columns[language] = index
            elif "SubjectCode" in value:
                # This column contains the mapped Subject Code
                self.subject_code_index = index
            elif "Wikidata" in value:
                # This column contains the Wikidata entity code
                self.wikidata_index = index

            index += 1

    def load_cv_items(self, existing_topics: Dict[str, CVItem]):
        """Loads the list of IPTC MediaTopic items from the sheet rows"""

        self.rows = []
        self._parent_qcode_stack = deque()
        self._parent_qcode_index = 0
        for row in self.worksheet.iter_rows(min_row=3):
            self.rows.append(self.convert_row_to_cv([
                cell.value
                for cell in row
            ], existing_topics))

    def convert_row_to_cv(self, row: List[str], existing_topics: Dict[str, CVItem]) -> CVItemFromIPTC:
        """Converts a MediaTopic from IPTC xlsx format to Superdesk CV format"""

        qcode = row[1][7:].strip()  # remove ``medtop:`` from the qcode
        original = existing_topics.get(qcode)
        parent = self._get_parent_qcode_for_row(row, qcode)

        # Get the column indices for the different language translations
        norwegian_index = self.name_columns.get("no")
        en_us_index = self.name_columns.get("en-us")
        en_gb_index = self.name_columns.get("en-gb")

        def get_name() -> str:
            if norwegian_index and row[norwegian_index]:
                return row[norwegian_index]
            elif en_us_index and row[en_us_index]:
                return row[en_us_index]
            elif en_gb_index and row[en_gb_index]:
                return row[en_gb_index]

            # This should not happen, every row should have at least one of the above
            return ""

        return CVItemFromIPTC(
            qcode=qcode,
            name=get_name().strip(),
            parent=parent.strip() if parent else None,
            iptc_subject=extract_code_from_string(row[self.subject_code_index]),
            wikidata=extract_code_from_string(row[self.wikidata_index]),
            is_active=original.get("is_active", True) if original else True,
            _existing=original,
            _missing_translation=not (norwegian_index and row[norwegian_index])
        )

    def _get_parent_qcode_for_row(self, row: List[str], qcode: str):
        """Get the parent qcode for the provided row

        :param list[str] row: The current row
        :param str qcode: The qcode of the current row

        The way parent qcodes are defined in the xlsx spreadsheet is not straightforward.
        We need to keep track of the parent level of the previous row(s) in order to determine
        the parent qcode of the current row.

        Example rows:

        =================== ====================== ================ ================ ================
        NewsCode-URI        NewsCode-QCode (flat)  Level1           Level2           Level3
        =================== ====================== ================ ================ ================
        mediatopic/01000000 medtop:01000000        medtop:01000000
        mediatopic/20000002 medtop:20000002                         medtop:20000002
        mediatopic/20000003 medtop:20000003                                          medtop:20000003
        mediatopic/20000038 medtop:20000038                         medtop:20000038
        mediatopic/20000045 medtop:20000045                         medtop:20000045
        mediatopic/20000051 medtop:20000051                                          medtop:20000051
        mediatopic/16000000 medtop:16000000        medtop:16000000
        =================== ====================== ================ ================ ================

        The cells in the different levels contain the qcode of that row (weird).
        Visually we can see the hierarchy of these topics, but programmatically is not straightforward.
        We use a stack to keep track of the current qcode and level as we iterate the rows.
        This allows us to calculate what the parent qcode of the current row is.
        """

        # Calculate the Level of this current row
        # By iterating over all the Level* columns (Level1 -> Level5)
        # The first column with a value is the level this row is at
        value_index: int = -1
        for index in range(2, 8):
            if row[index]:
                value_index = index - 2
                break

        next_parent: Optional[str] = None
        if value_index == -1:
            # Every row should have an entry in one of the level columns
            # If there is none, we'll have to assume there is no parent for this item
            return None
        elif value_index == 0:
            # We're at Level1, meaning this row has no parent
            # So we can reset the qcode stack/index
            self._parent_qcode_stack.clear()
            self._parent_qcode_index = 0

            # Append qcode of this current row
            # (in case the next row is a child of this one)
            self._parent_qcode_stack.append(qcode)
        elif value_index == self._parent_qcode_index:
            # The level has not changed from the previous row

            # Pop the qcode from the stack (qcode from the previous row)
            self._parent_qcode_stack.pop()

            # Grab the qcode of our parent
            next_parent = self._parent_qcode_stack[-1]

            # Append the qcode of this current row
            # (in case the next row is a child of this one)
            self._parent_qcode_stack.append(qcode)
        elif value_index > self._parent_qcode_index:
            # The level has increased, which means this row is a child of the previous row

            # Grab the qcode of the previous row
            next_parent = self._parent_qcode_stack[-1]

            # Append the qcode of this current row
            # (in case the next row is a child of this one)
            self._parent_qcode_stack.append(qcode)

            # Set the new level, to the current row's level
            self._parent_qcode_index = value_index
        elif value_index < self._parent_qcode_index:
            # The level has decreased (possibly more than 1 leven).

            # Pop the previous row's qcode from the stack
            self._parent_qcode_stack.pop()

            # Calculate how many level's we have gone back
            # and pop those qcodes from the stack
            for _ in range(0, self._parent_qcode_index - value_index):
                self._parent_qcode_stack.pop()

            try:
                # Attempt to grab the qcode from the end of the stack
                # to obtain the new parent for this current row
                next_parent = self._parent_qcode_stack[-1]
                self._parent_qcode_index = value_index
            except IndexError:
                # If an ``IndexError`` is raised, that means we're back at Level1.
                # So we can reset the qcode stack/index
                self._parent_qcode_stack.clear()
                self._parent_qcode_index = 0

            # Finally, append the qcode of this current row
            # (in case the next row is a child of this one)
            self._parent_qcode_stack.append(qcode)

        # Return the calculated qcode of the parent to the current row
        # (if one was found)
        return next_parent
